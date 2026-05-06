import streamlit as st
import psycopg2
import pandas as pd
import calendar
import requests
import io
from datetime import datetime, date, timedelta
from calendar import monthrange
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURAÇÃO DA PÁGINA WEB
# ==========================================
st.set_page_config(page_title="Gestão de Escala e Férias", page_icon="📅", layout="wide")
calendar.setfirstweekday(calendar.SUNDAY)

# ==========================================
# SEGURANÇA: PUXANDO CREDENCIAIS DO SECRETS
# ==========================================
try:
    db = st.secrets["connections"]["postgresql"]
    DB_URI = f"postgresql://{db['user']}:{db['password']}@{db['host']}:{db['port']}/{db['database']}"
except KeyError:
    st.error("🚨 As credenciais do banco de dados não foram encontradas no st.secrets.")
    st.stop()

# ==========================================
# CONEXÃO COM O BANCO DE DADOS (PostgreSQL)
# ==========================================
@st.cache_resource
def init_connection():
    return psycopg2.connect(DB_URI)

try:
    conn = init_connection()
    conn.autocommit = True
    cursor = conn.cursor()
    
    # --- AUTO-ATUALIZAÇÃO DO BANCO DE DADOS ---
    cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name='colaboradores' AND column_name='exibir_escala';")
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE colaboradores ADD COLUMN exibir_escala INTEGER DEFAULT 1;")
    # ------------------------------------------
except Exception as e:
    st.error(f"🚨 Erro de Conexão com o Supabase. Detalhe: {e}")
    st.stop()

# ==========================================
# SISTEMA DE LOGIN (SEGURANÇA DA NUVEM)
# ==========================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.title("🔒 Acesso Restrito - TI Camarotti")
    st.write("Por favor, insira a senha corporativa para gerenciar as escalas.")
    senha = st.text_input("Senha de Acesso", type="password")
    if st.button("Entrar"):
        if senha == st.secrets["senha_acesso"]: 
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Senha incorreta!")
    st.stop()

# ==========================================
# MOTORES DE CÁLCULO E LÓGICA
# ==========================================
def get_saldos(colab_id):
    cursor.execute("SELECT dias_pendentes, saldo_bh FROM colaboradores WHERE id=%s", (colab_id,))
    res = cursor.fetchone()
    if not res: return 0, 0.0, 0, 0.0, 0, 0, 0.0
    
    dias_db = int(res[0])
    bh_db = float(res[1])
    hoje = datetime.now()
    f_efetivas, f_oficial, f_bh = 0, 0, 0.0

    cursor.execute("SELECT tipo, data_inicio, numero_dias, valor_descontado FROM lancamentos WHERE colaborador_id=%s", (colab_id,))
    for tipo, d_ini_str, num_dias, v_desc in cursor.fetchall():
        try:
            dt_ini = datetime.strptime(d_ini_str, "%d/%m/%Y")
            if dt_ini.year > hoje.year or (dt_ini.year == hoje.year and dt_ini.month > hoje.month):
                if tipo == "Férias (Efetivas)": f_efetivas += num_dias
                elif tipo == "Férias (Oficial)": f_oficial += num_dias
                elif tipo == "Folga BH": f_bh += float(v_desc)
        except: pass

    dias_atual = dias_db + f_efetivas - f_oficial
    bh_atual = bh_db + f_bh
    return dias_atual, bh_atual, dias_db, bh_db, f_efetivas, f_oficial, f_bh

def obter_datas_ocupadas(colaborador_id):
    cursor.execute("SELECT data_inicio, data_fim FROM lancamentos WHERE colaborador_id = %s AND tipo != 'Férias (Oficial)'", (colaborador_id,))
    lancamentos = cursor.fetchall()
    datas_ocupadas = set()
    for d_ini_str, d_fim_str in lancamentos:
        try:
            d_ini = datetime.strptime(d_ini_str, "%d/%m/%Y")
            d_fim = datetime.strptime(d_fim_str, "%d/%m/%Y")
            atual = d_ini
            while atual <= d_fim:
                datas_ocupadas.add(atual.strftime("%d/%m/%Y"))
                atual += timedelta(days=1)
        except: pass
    return datas_ocupadas

# ==========================================
# MENU DE NAVEGAÇÃO LATERAL
# ==========================================
st.sidebar.title("📅 Gestão de Escala")
st.sidebar.markdown("---")
menu = st.sidebar.radio("Selecione o Módulo:", [
    "📊 Dashboard Interativo", 
    "👥 Gestão de Equipe", 
    "✈️ Lançamentos",
    "🌴 Gerir Feriados"
])

# ==========================================
# MÓDULO 1: DASHBOARD
# ==========================================
if menu == "📊 Dashboard Interativo":
    st.header("Dashboard da Equipe")
    
    # ----------------------------------------------------
    # SISTEMA DE ALERTA DP
    # ----------------------------------------------------
    cursor.execute("SELECT id, nome, venc_ferias FROM colaboradores WHERE ativo = 1")
    alertas_aviso = []
    alertas_vencidos = []
    hoje_date = date.today()
    
    for colab_id, nome, venc_str in cursor.fetchall():
        if venc_str:
            try:
                dt_limite = datetime.strptime(venc_str, "%d/%m/%Y").date()
                dt_prazo_gestor = dt_limite - timedelta(days=45)
                dif_dias = (dt_prazo_gestor - hoje_date).days
                
                cursor.execute("SELECT data_inicio FROM lancamentos WHERE colaborador_id=%s AND tipo LIKE 'Férias%%'", (colab_id,))
                ja_resolveu = False
                for (d_ini_str,) in cursor.fetchall():
                    try:
                        dt_ini_lancamento = datetime.strptime(d_ini_str, "%d/%m/%Y").date()
                        if dt_ini_lancamento >= (dt_limite - timedelta(days=365)):
                            ja_resolveu = True
                            break
                    except:
                        pass
                
                if not ja_resolveu:
                    if 0 <= dif_dias <= 45:
                        alertas_aviso.append(f"⚠️ **{nome}**: Enviar pedido de férias ao DP até **{dt_prazo_gestor.strftime('%d/%m/%Y')}** (Prazo vence em {dif_dias} dias!)")
                    elif dif_dias < 0:
                        alertas_vencidos.append(f"🚨 **{nome}**: Prazo de envio ao DP **VENCIDO**! (Era até {dt_prazo_gestor.strftime('%d/%m/%Y')})")
            except:
                pass
    
    if alertas_vencidos or alertas_aviso:
        with st.container():
            for a in alertas_vencidos:
                st.error(a)
            for a in alertas_aviso:
                st.warning(a)
        st.markdown("<br>", unsafe_allow_html=True)

    # ----------------------------------------------------
    # FILTROS PRINCIPAIS
    # ----------------------------------------------------
    col1, col2 = st.columns([1, 3])
    ano_atual = datetime.now().year
    mes_atual = datetime.now().month
    
    with col1:
        ano_selecionado = st.selectbox("Ano", range(ano_atual - 2, ano_atual + 3), index=2)
    with col2:
        meses = ["Resumo Anual", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        mes_selecionado = st.selectbox("Mês / Visão", meses, index=mes_atual)
    
    st.markdown("**Filtros de Presença:**")
    col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
    mostrar_presencial = col_f1.checkbox("🟩 Presencial", value=True)
    mostrar_efetivas = col_f2.checkbox("🟧 Férias Efetivas", value=True)
    mostrar_oficial = col_f3.checkbox("🟪 Férias Oficial", value=False)
    mostrar_folga_bh = col_f4.checkbox("🟦 Folga BH", value=True)
    mostrar_feriados = col_f5.checkbox("🟫 Feriados", value=True)
    
    st.markdown("---")
    
    # ----------------------------------------------------
    # BUSCA DE DADOS E FERIADOS
    # ----------------------------------------------------
    cursor.execute("SELECT c.nome, l.tipo, l.data_inicio, l.data_fim FROM lancamentos l JOIN colaboradores c ON l.colaborador_id = c.id WHERE c.ativo = 1 AND c.exibir_escala = 1")
    todos_lancamentos = cursor.fetchall()
    
    cursor.execute("SELECT data_feriado, descricao FROM feriados")
    feriados = {linha[0]: linha[1] for linha in cursor.fetchall()}

    # ----------------------------------------------------
    # RENDERIZAÇÃO DA VISÃO "RESUMO ANUAL"
    # ----------------------------------------------------
    if mes_selecionado == "Resumo Anual":
        st.subheader(f"Resumo de Férias do Ano - {ano_selecionado}")
        dados_resumo = []
        for nome, tipo, d_ini, d_fim in todos_lancamentos:
            if "Férias" in tipo:
                try:
                    dt_ini = datetime.strptime(d_ini, "%d/%m/%Y")
                    dt_fim = datetime.strptime(d_fim, "%d/%m/%Y")
                    if dt_ini.year == ano_selecionado or dt_fim.year == ano_selecionado:
                        dias = (dt_fim - dt_ini).days + 1
                        dados_resumo.append({"Colaborador": nome, "Tipo": tipo, "Início": d_ini, "Fim": d_fim, "Dias": dias, "SortDate": dt_ini})
                except: pass
        
        if dados_resumo:
            df_resumo = pd.DataFrame(dados_resumo).sort_values(by="SortDate").drop(columns=["SortDate"])
            
            c_res1, c_res2 = st.columns([5, 1])
            with c_res1:
                st.dataframe(df_resumo, use_container_width=True)
            with c_res2:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_resumo.to_excel(writer, index=False, sheet_name=f'Férias_{ano_selecionado}')
                    worksheet = writer.sheets[f'Férias_{ano_selecionado}']
                    
                    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    for col in worksheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(cell.value)
                            except: pass
                        worksheet.column_dimensions[column].width = max_length + 2
                
                st.download_button(
                    label="📥 Exportar Excel",
                    data=output.getvalue(),
                    file_name=f'Resumo_Ferias_{ano_selecionado}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )
        else:
            st.info("Nenhuma férias programada para este ano.")

    # ----------------------------------------------------
    # RENDERIZAÇÃO DA VISÃO "MENSAL" (Calendário ou Matriz)
    # ----------------------------------------------------
    else:
        mes_num = meses.index(mes_selecionado)
        cal = calendar.monthcalendar(ano_selecionado, mes_num)
        dias_no_mes = calendar.monthrange(ano_selecionado, mes_num)[1]
        
        eventos_mes = {d: [] for d in range(1, dias_no_mes + 1)}
        for nome, tipo, d_ini_str, d_fim_str in todos_lancamentos:
            try:
                dt_ini = datetime.strptime(d_ini_str, "%d/%m/%Y").date()
                dt_fim = datetime.strptime(d_fim_str, "%d/%m/%Y").date()
                atual = dt_ini
                while atual <= dt_fim:
                    if atual.month == mes_num and atual.year == ano_selecionado:
                        eventos_mes[atual.day].append((nome, tipo))
                    atual += timedelta(days=1)
            except: pass
            
        cursor.execute("SELECT nome FROM colaboradores WHERE ativo = 1 AND exibir_escala = 1 ORDER BY nome ASC")
        todos_colabs_export = [row[0] for row in cursor.fetchall()]
        
        # --- GERADOR DO EXCEL DE DOWNLOAD (EM BACKGROUND) ---
        matriz_dados = []
        row_semana = {"Colaborador": "Dia da Semana ➔"}
        for dia in range(1, dias_no_mes + 1):
            wd = calendar.weekday(ano_selecionado, mes_num, dia)
            row_semana[str(dia)] = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][wd]
        matriz_dados.append(row_semana)

        for nome_colab in todos_colabs_export:
            row = {"Colaborador": nome_colab}
            for dia in range(1, dias_no_mes + 1): row[str(dia)] = ""
            matriz_dados.append(row)

        df_export = pd.DataFrame(matriz_dados)
        df_export.set_index("Colaborador", inplace=True)
        siglas_export = {"Presencial": "P", "Férias (Efetivas)": "FE", "Férias (Oficial)": "FO", "Folga BH": "BH"}

        for dia, eventos in eventos_mes.items():
            for ev_nome_full, ev_tipo in eventos:
                if ev_tipo == "Presencial" and not mostrar_presencial: continue
                if ev_tipo == "Férias (Efetivas)" and not mostrar_efetivas: continue
                if ev_tipo == "Férias (Oficial)" and not mostrar_oficial: continue
                if ev_tipo == "Folga BH" and not mostrar_folga_bh: continue
                if ev_nome_full in df_export.index:
                    df_export.at[ev_nome_full, str(dia)] = siglas_export.get(ev_tipo, ev_tipo)

        df_export.reset_index(inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f'{mes_selecionado}_{ano_selecionado}'
            df_export.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = 'B3'
            
            thin_border = Border(left=Side(style='thin', color='D4D4D4'), right=Side(style='thin', color='D4D4D4'), top=Side(style='thin', color='D4D4D4'), bottom=Side(style='thin', color='D4D4D4'))
            fill_fds = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
            fill_header = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
            fill_semana = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            fill_feriado = PatternFill(start_color="FFF0CC", end_color="FFF0CC", fill_type="solid")
            
            fill_p = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
            fill_fe = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
            fill_fo = PatternFill(start_color="7E22CE", end_color="7E22CE", fill_type="solid")
            fill_bh = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
            
            font_header = Font(color="FFFFFF", bold=True)
            font_semana = Font(color="005CE6", bold=True, size=9)
            font_feriado = Font(color="B45309", bold=True, size=9)
            font_branca = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            
            worksheet.column_dimensions['A'].width = 36
            
            for row_idx in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = 20
                for col_idx in range(1, dias_no_mes + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if col_idx > 1: cell.alignment = align_center
            
            for cell in worksheet[1]:
                cell.fill = fill_header
                cell.font = font_header

            worksheet.cell(row=2, column=1).alignment = Alignment(horizontal="right", vertical="center")
            worksheet.cell(row=2, column=1).font = font_semana
            
            for col_idx in range(2, dias_no_mes + 2):
                dia_num = col_idx - 1
                data_str = f"{dia_num:02d}/{mes_num:02d}/{ano_selecionado}"
                is_feriado = data_str in feriados
                cell_semana = worksheet.cell(row=2, column=col_idx)
                if is_feriado:
                    cell_semana.fill = fill_feriado
                    cell_semana.font = font_feriado
                else:
                    cell_semana.fill = fill_semana
                    cell_semana.font = font_semana

            for col_idx in range(2, dias_no_mes + 2):
                dia_num = col_idx - 1
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 6
                
                data_str = f"{dia_num:02d}/{mes_num:02d}/{ano_selecionado}"
                is_feriado = data_str in feriados
                dia_semana = calendar.weekday(ano_selecionado, mes_num, dia_num)
                is_weekend = (dia_semana == 5 or dia_semana == 6)
                
                for row_idx in range(3, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if is_feriado: cell.fill = fill_feriado
                    elif is_weekend: cell.fill = fill_fds

                    val = cell.value
                    if val == "P": cell.fill = fill_p; cell.font = font_branca
                    elif val == "FE": cell.fill = fill_fe; cell.font = font_branca
                    elif val == "FO": cell.fill = fill_fo; cell.font = font_branca
                    elif val == "BH": cell.fill = fill_bh; cell.font = font_branca
                    
            start_leg = worksheet.max_row + 2
            worksheet.cell(row=start_leg, column=1, value="LEGENDA DE CORES:").font = Font(bold=True, color="555555")
            legendas = [("Presencial", "P", fill_p, font_branca), ("Férias Efetivas", "FE", fill_fe, font_branca), ("Férias Oficial", "FO", fill_fo, font_branca), ("Folga Banco de Horas", "BH", fill_bh, font_branca), ("Feriado Nacional", "", fill_feriado, None), ("Fim de Semana", "", fill_fds, None)]
            for i, (texto, sigla, fill, font) in enumerate(legendas):
                row_leg = start_leg + 1 + i
                worksheet.row_dimensions[row_leg].height = 20
                c_text = worksheet.cell(row=row_leg, column=1, value=texto)
                c_text.alignment = Alignment(horizontal="right", vertical="center")
                c_text.font = Font(italic=True, color="555555")
                c_box = worksheet.cell(row=row_leg, column=2, value=sigla)
                c_box.fill = fill
                c_box.border = thin_border
                c_box.alignment = align_center
                if font: c_box.font = font

        # --- SELETOR DE VISUALIZAÇÃO DA WEB E EXPORTAÇÃO ---
        c_view1, c_view2 = st.columns([3, 1])
        with c_view1:
            visao_web = st.radio("Formato de Exibição Web:", ["📊 Matriz de Escala (Heatmap)", "🗓️ Calendário Clássico"], horizontal=True)
        with c_view2:
            st.download_button(
                label="📥 Exportar Matriz (Excel)",
                data=output.getvalue(),
                file_name=f'Escala_Matriz_{mes_selecionado}_{ano_selecionado}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True
            )

        st.markdown("<br>", unsafe_allow_html=True)

        cores = {"Presencial": "#15803d", "Férias (Efetivas)": "#c2410c", "Férias (Oficial)": "#7e22ce", "Folga BH": "#1d4ed8"}

        # --- OPÇÃO 1: MATRIZ NA WEB ---
        if visao_web == "📊 Matriz de Escala (Heatmap)":
            html_matrix = "<div style='overflow-x: auto;'>"
            html_matrix += "<table style='width:100%; border-collapse: collapse; text-align:center; font-family:sans-serif; font-size: 13px; white-space: nowrap;'>"
            
            # CABEÇALHO (DIAS)
            html_matrix += "<tr>"
            html_matrix += "<th style='position: sticky; left: 0; background-color: #1f538d; color: white; padding: 8px; text-align: left; z-index: 2; min-width: 200px; border: 1px solid #ddd;'>Colaborador</th>"
            for dia in range(1, dias_no_mes + 1):
                html_matrix += f"<th style='background-color:#1f538d; color:white; padding: 8px; border: 1px solid #ddd; min-width: 35px;'>{dia}</th>"
            html_matrix += "</tr>"
            
            # DIA DA SEMANA
            html_matrix += "<tr style='background-color:#E6F2FF; color:#005CE6; font-size: 11px;'>"
            html_matrix += "<td style='position: sticky; left: 0; background-color: #E6F2FF; padding: 4px; text-align: right; font-weight: bold; border: 1px solid #ddd; z-index: 1;'>Dia da Semana ➔</td>"
            for dia in range(1, dias_no_mes + 1):
                wd = calendar.weekday(ano_selecionado, mes_num, dia)
                sigla_dia = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][wd]
                
                data_str = f"{dia:02d}/{mes_num:02d}/{ano_selecionado}"
                if data_str in feriados and mostrar_feriados:
                    html_matrix += f"<td style='background-color:#FFF0CC; color:#B45309; padding: 4px; border: 1px solid #ddd; font-weight: bold;'>{sigla_dia}</td>"
                else:
                    html_matrix += f"<td style='padding: 4px; border: 1px solid #ddd; font-weight: bold;'>{sigla_dia}</td>"
            html_matrix += "</tr>"

            # COLABORADORES
            for nome_colab in todos_colabs_export:
                html_matrix += "<tr>"
                html_matrix += f"<td style='position: sticky; left: 0; background-color: white; padding: 8px; text-align: left; font-weight: bold; border: 1px solid #ddd; z-index: 1; color: #333;'>{nome_colab}</td>"
                
                for dia in range(1, dias_no_mes + 1):
                    data_str = f"{dia:02d}/{mes_num:02d}/{ano_selecionado}"
                    dia_semana = calendar.weekday(ano_selecionado, mes_num, dia)
                    is_weekend = (dia_semana == 5 or dia_semana == 6)
                    is_feriado = data_str in feriados
                    
                    bg_color = "#ffffff"
                    if is_feriado and mostrar_feriados: bg_color = "#FFF0CC"
                    elif is_weekend: bg_color = "#f0f2f6"
                    
                    txt_sigla = ""
                    txt_color = ""
                    
                    # Procura o evento deste colaborador neste dia
                    for ev_nome_full, ev_tipo in eventos_mes[dia]:
                        if ev_nome_full == nome_colab:
                            if ev_tipo == "Presencial" and not mostrar_presencial: continue
                            if ev_tipo == "Férias (Efetivas)" and not mostrar_efetivas: continue
                            if ev_tipo == "Férias (Oficial)" and not mostrar_oficial: continue
                            if ev_tipo == "Folga BH" and not mostrar_folga_bh: continue
                            
                            bg_color = cores.get(ev_tipo, bg_color)
                            txt_color = "white"
                            txt_sigla = siglas_export.get(ev_tipo, "")
                            break
                    
                    if txt_sigla:
                        html_matrix += f"<td style='background-color:{bg_color}; color:{txt_color}; padding: 8px; border: 1px solid #ddd; font-weight: bold;'>{txt_sigla}</td>"
                    else:
                        html_matrix += f"<td style='background-color:{bg_color}; padding: 8px; border: 1px solid #ddd;'></td>"
                        
                html_matrix += "</tr>"
            
            html_matrix += "</table></div>"
            st.markdown(html_matrix, unsafe_allow_html=True)

            # Legenda Minimalista abaixo da Tabela Web
            st.markdown("""
            <div style='margin-top: 15px; font-size: 13px; color: #555;'>
                <b>Legenda:</b> 
                <span style='background-color: #15803D; color: white; padding: 2px 6px; border-radius: 4px; margin-left: 5px;'>P</span> Presencial 
                <span style='background-color: #C2410C; color: white; padding: 2px 6px; border-radius: 4px; margin-left: 10px;'>FE</span> Férias Efetivas 
                <span style='background-color: #7E22CE; color: white; padding: 2px 6px; border-radius: 4px; margin-left: 10px;'>FO</span> Férias Oficial 
                <span style='background-color: #1D4ED8; color: white; padding: 2px 6px; border-radius: 4px; margin-left: 10px;'>BH</span> Folga BH
                <span style='background-color: #FFF0CC; color: #B45309; padding: 2px 6px; border-radius: 4px; margin-left: 10px; border: 1px solid #DDD;'>Dia</span> Feriado
                <span style='background-color: #f0f2f6; color: #005CE6; padding: 2px 6px; border-radius: 4px; margin-left: 10px; border: 1px solid #DDD;'>Dia</span> Fim de Semana
            </div>
            """, unsafe_allow_html=True)

        # --- OPÇÃO 2: CALENDÁRIO CLÁSSICO ---
        else:
            html_cal = "<table style='width:100%; border-collapse: collapse; text-align:center; font-family:sans-serif;'>"
            html_cal += "<tr style='background-color:#1f538d; color:white;'><th>Dom</th><th>Seg</th><th>Ter</th><th>Qua</th><th>Qui</th><th>Sex</th><th>Sáb</th></tr>"
            
            for semana in cal:
                html_cal += "<tr>"
                for col, dia in enumerate(semana):
                    if dia == 0:
                        html_cal += "<td style='border: 1px solid #ddd; background-color:#f0f2f6; padding:10px;'></td>"
                    else:
                        data_str = f"{dia:02d}/{mes_num:02d}/{ano_selecionado}"
                        bg_color = "#ffffff" if (col != 0 and col != 6) else "#f9f9f9"
                        
                        eventos_html = ""
                        if data_str in feriados and mostrar_feriados:
                            bg_color = "#fff0cc"
                            eventos_html += f"<div style='color:#b45309; font-size:12px; font-weight:bold; margin-bottom:2px;'>★ {feriados[data_str]}</div>"
                        
                        for ev_nome_full, ev_tipo in eventos_mes[dia]:
                            if ev_tipo == "Presencial" and not mostrar_presencial: continue
                            if ev_tipo == "Férias (Efetivas)" and not mostrar_efetivas: continue
                            if ev_tipo == "Férias (Oficial)" and not mostrar_oficial: continue
                            if ev_tipo == "Folga BH" and not mostrar_folga_bh: continue

                            cor_txt = cores.get(ev_tipo, "#000")
                            ev_nome = ev_nome_full.split()[0]
                            eventos_html += f"<div style='color:{cor_txt}; font-size:12px; font-weight:bold; margin-top:2px;'>■ {ev_nome}</div>"
                        
                        html_cal += f"<td style='border: 1px solid #ddd; background-color:{bg_color}; padding:10px; vertical-align:top; min-height:80px; width:14%;'>"
                        html_cal += f"<div style='font-size:16px; font-weight:bold; color:#555; text-align:right;'>{dia}</div>"
                        html_cal += eventos_html
                        html_cal += "</td>"
                html_cal += "</tr>"
            html_cal += "</table>"
            st.markdown(html_cal, unsafe_allow_html=True)

# ==========================================
# MÓDULO 2: GESTÃO DE EQUIPE
# ==========================================
elif menu == "👥 Gestão de Equipe":
    st.header("Gestão de Colaboradores")
    
    cursor.execute("SELECT id, nome, ativo FROM colaboradores ORDER BY ativo DESC, nome ASC")
    colaboradores = cursor.fetchall()
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("Equipe")
        lista_nomes = ["➕ Cadastrar Novo"] + [f"{c[1]} {'(Inativo)' if c[2]==0 else ''}" for c in colaboradores]
        escolha = st.radio("Selecione:", lista_nomes)
    
    with col2:
        if escolha == "➕ Cadastrar Novo":
            st.subheader("Novo Colaborador")
            with st.form("form_novo_colab"):
                nome = st.text_input("Nome Completo")
                funcao = st.text_input("Função")
                area = st.text_input("Área")
                admissao = st.date_input("Data de Admissão", format="DD/MM/YYYY")
                data_limite = st.date_input("Data Limite (Baseada na Planilha DP)", format="DD/MM/YYYY")
                d_pendentes = st.number_input("Dias Pendentes (Saldo Inicial)", value=0, step=1)
                bh_inicial = st.number_input("Saldo Banco de Horas Inicial", value=0.0, step=0.5)
                
                st.markdown("<br>", unsafe_allow_html=True)
                exibir_escala = st.checkbox("👁️ Exibir Colaborador na Grade/Dashboard?", value=True)
                val_exibir = 1 if exibir_escala else 0
                
                if st.form_submit_button("💾 Salvar Colaborador", use_container_width=True):
                    try:
                        cursor.execute("INSERT INTO colaboradores (nome, funcao, area, admissao, venc_ferias, dias_pendentes, saldo_bh, exibir_escala) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", 
                                       (nome, funcao, area, admissao.strftime("%d/%m/%Y"), data_limite.strftime("%d/%m/%Y"), d_pendentes, bh_inicial, val_exibir))
                        st.success("Colaborador cadastrado!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro: {e}")
        else:
            idx = lista_nomes.index(escolha) - 1
            colab_id = colaboradores[idx][0]
            
            cursor.execute("SELECT nome, funcao, area, admissao, venc_ferias, ativo, exibir_escala FROM colaboradores WHERE id=%s", (colab_id,))
            dados = cursor.fetchone()
            d_atual, b_atual, d_db, b_db, f_efe, f_ofic, f_bh = get_saldos(colab_id)
            
            exibir_bd = dados[6] if dados[6] is not None else 1
            
            c_h1, c_h2, c_h3 = st.columns([2, 1, 1])
            c_h1.subheader(dados[0])
            c_h2.markdown(f"<div style='background-color:#fff2e6; color:#cc6600; padding:10px; border-radius:5px; text-align:center; font-weight:bold; border: 1px solid #ffcc99;'>✈ Dias Pendentes: {d_atual}</div>", unsafe_allow_html=True)
            c_h3.markdown(f"<div style='background-color:#e6f2ff; color:#005ce6; padding:10px; border-radius:5px; text-align:center; font-weight:bold; border: 1px solid #99c2ff;'>⏳ Banco de Horas: {b_atual}h</div>", unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            if f_efe > 0 or f_ofic > 0 or f_bh > 0:
                msg_futuro = "📅 **Agendado para meses futuros:**"
                if f_ofic > 0: msg_futuro += f" &nbsp; `+{f_ofic}d (Assinar)`"
                if f_efe > 0: msg_futuro += f" &nbsp; `-{f_efe}d (Efetivas)`"
                if f_bh > 0: msg_futuro += f" &nbsp; `-{f_bh}h (Folga)`"
                st.warning(msg_futuro)

            with st.form("form_edita_colab"):
                nome = st.text_input("Nome Completo", value=dados[0])
                funcao = st.text_input("Função", value=dados[1])
                area = st.text_input("Área", value=dados[2])
                
                try:
                    data_adm_obj = datetime.strptime(dados[3], "%d/%m/%Y").date()
                except:
                    try:
                        data_adm_obj = datetime.strptime(dados[3], "%Y-%m-%d").date()
                    except:
                        data_adm_obj = date.today()
                
                try:
                    data_limite_obj = datetime.strptime(dados[4], "%d/%m/%Y").date()
                except:
                    data_limite_obj = date.today()
                    
                admissao = st.date_input("Data de Admissão", value=data_adm_obj, format="DD/MM/YYYY")
                data_limite = st.date_input("Data Limite (Baseada na Planilha DP)", value=data_limite_obj, format="DD/MM/YYYY")
                
                d_pendentes = st.number_input("Dias Pendentes (Mês Atual)", value=int(d_atual), step=1)
                bh_atual_input = st.number_input("Saldo Banco de Horas (Mês Atual)", value=float(b_atual), step=0.5)
                
                st.markdown("<br>", unsafe_allow_html=True)
                exibir_escala = st.checkbox("👁️ Exibir Colaborador na Grade/Dashboard?", value=bool(exibir_bd))
                val_exibir = 1 if exibir_escala else 0
                
                st.markdown("<br>", unsafe_allow_html=True)
                col_btn1, col_btn2, col_btn3 = st.columns(3)
                
                with col_btn1:
                    submit = st.form_submit_button("💾 Atualizar", type="primary", use_container_width=True)
                with col_btn2:
                    btn_status = "Arquivar" if dados[5] == 1 else "Reativar"
                    status_submit = st.form_submit_button(f"📦 {btn_status}", use_container_width=True)
                with col_btn3:
                    excluir_submit = st.form_submit_button("🗑️ Excluir Definitivo", use_container_width=True)

                if submit:
                    novo_d_db = d_pendentes - f_efe + f_ofic
                    novo_b_db = bh_atual_input - f_bh
                    cursor.execute("UPDATE colaboradores SET nome=%s, funcao=%s, area=%s, admissao=%s, venc_ferias=%s, dias_pendentes=%s, saldo_bh=%s, exibir_escala=%s WHERE id=%s", 
                                   (nome, funcao, area, admissao.strftime("%d/%m/%Y"), data_limite.strftime("%d/%m/%Y"), novo_d_db, novo_b_db, val_exibir, colab_id))
                    st.success("Atualizado!")
                    st.rerun()
                
                if status_submit:
                    novo_st = 0 if dados[5] == 1 else 1
                    cursor.execute("UPDATE colaboradores SET ativo=%s WHERE id=%s", (novo_st, colab_id))
                    st.rerun()

                if excluir_submit:
                    cursor.execute("DELETE FROM lancamentos WHERE colaborador_id=%s", (colab_id,))
                    cursor.execute("DELETE FROM colaboradores WHERE id=%s", (colab_id,))
                    st.rerun()

# ==========================================
# MÓDULO 3: LANÇAMENTOS 
# ==========================================
elif menu == "✈️ Lançamentos":
    st.header("Lançamentos Individuais")
    
    cursor.execute("SELECT id, nome FROM colaboradores WHERE ativo = 1 ORDER BY nome ASC")
    colabs = cursor.fetchall()
    
    if not colabs:
        st.warning("Cadastre um colaborador primeiro.")
        st.stop()
        
    colab_selecionado = st.selectbox("Selecione o Colaborador:", [c[1] for c in colabs])
    colab_id = [c[0] for c in colabs if c[1] == colab_selecionado][0]
    
    d_atual, b_atual, d_db, b_db, _, _, _ = get_saldos(colab_id)
    
    st.markdown(f"**Saldo Atual:** `{d_atual} dias` | Banco de Horas: `{b_atual}h`")
    
    tipo_lancamento = st.radio("O que deseja lançar?", ["Férias", "Presencial", "Folga BH"], horizontal=True)
    
    # ----------------------------------------------------
    # LÓGICA 1: FÉRIAS
    # ----------------------------------------------------
    if tipo_lancamento == "Férias":
        with st.form("form_ferias"):
            acao_ferias = st.selectbox("Ação de Férias:", ["✍️ Assinar (Férias Oficial) -> ADICIONA Saldo", "🏖️ Efetivar (Férias Efetivas) -> DESCONTA Saldo"])
            tipo_bd = "Férias (Oficial)" if "Oficial" in acao_ferias else "Férias (Efetivas)"
            
            col1, col2 = st.columns(2)
            d_inicio = col1.date_input("Data de Início", format="DD/MM/YYYY")
            qtd_dias_input = col2.number_input("Quantidade de Dias", value=20, min_value=1, step=1)
            
            btn_lancar = st.form_submit_button("🚀 Gravar Férias", type="primary", use_container_width=True)
            
            if btn_lancar:
                d_fim = d_inicio + timedelta(days=qtd_dias_input - 1)
                qtd_dias = qtd_dias_input
                
                if tipo_bd == "Férias (Efetivas)" and (d_db - qtd_dias) < 0:
                    st.error(f"Saldo final insuficiente! Projetado: {d_db} dias. Tentativa: {qtd_dias} dias.")
                else:
                    cursor.execute("INSERT INTO lancamentos (colaborador_id, tipo, data_inicio, data_fim, numero_dias, valor_descontado) VALUES (%s, %s, %s, %s, %s, %s)",
                                   (colab_id, tipo_bd, d_inicio.strftime("%d/%m/%Y"), d_fim.strftime("%d/%m/%Y"), qtd_dias, 0))
                    
                    if tipo_bd == "Férias (Oficial)":
                        cursor.execute("UPDATE colaboradores SET dias_pendentes = dias_pendentes + %s WHERE id = %s", (qtd_dias, colab_id))
                    elif tipo_bd == "Férias (Efetivas)":
                        cursor.execute("UPDATE colaboradores SET dias_pendentes = dias_pendentes - %s WHERE id = %s", (qtd_dias, colab_id))
                    
                    st.success(f"Férias lançadas com sucesso! Fim calculado para: {d_fim.strftime('%d/%m/%Y')}")
                    st.rerun()

    # ----------------------------------------------------
    # LÓGICA 2: PRESENCIAL E FOLGA (Dias Picados via Lote)
    # ----------------------------------------------------
    else:
        st.markdown("---")
        st.markdown(f"### 📅 Selecionar Dias - {tipo_lancamento}")
        
        if tipo_lancamento == "Folga BH":
            horas_abater = st.number_input("Horas a abater por dia:", value=8.0, step=0.5)
            tipo_bd = "Folga BH"
        else:
            horas_abater = 0
            tipo_bd = "Presencial"
            
        col_m, col_a = st.columns(2)
        meses_lista = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        mes_sel = col_m.selectbox("Mês Alvo", meses_lista, index=datetime.now().month - 1)
        ano_sel = col_a.number_input("Ano Alvo", value=datetime.now().year, step=1)
        
        mes_num = meses_lista.index(mes_sel) + 1
        cal_matriz = calendar.monthcalendar(ano_sel, mes_num)
        
        cursor.execute("SELECT id, data_inicio, valor_descontado FROM lancamentos WHERE colaborador_id=%s AND tipo=%s", (colab_id, tipo_bd))
        mapa_existentes = {}
        for l_id, d_ini_str, v_desc in cursor.fetchall():
            if d_ini_str.endswith(f"/{mes_num:02d}/{ano_sel}"):
                mapa_existentes[d_ini_str] = (l_id, v_desc)
        
        cursor.execute("SELECT data_inicio, data_fim, tipo FROM lancamentos WHERE colaborador_id=%s AND tipo!=%s", (colab_id, tipo_bd))
        outros_lancamentos = {}
        for d_ini_str, d_fim_str, t_bd in cursor.fetchall():
            if t_bd == "Férias (Oficial)":
                continue
                
            try:
                dt_ini = datetime.strptime(d_ini_str, "%d/%m/%Y")
                dt_fim = datetime.strptime(d_fim_str, "%d/%m/%Y")
                atual = dt_ini
                while atual <= dt_fim:
                    if atual.month == mes_num and atual.year == ano_sel:
                        outros_lancamentos[atual.strftime("%d/%m/%Y")] = t_bd
                    atual += timedelta(days=1)
            except:
                pass

        cursor.execute("SELECT data_feriado, descricao FROM feriados")
        feriados_lote = {linha[0]: linha[1] for linha in cursor.fetchall()}

        with st.form("form_dias_lote"):
            st.write("Marque as caixinhas dos dias que deseja lançar abaixo:")
            dias_semana = ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb"]
            cols_header = st.columns(7)
            for i, d in enumerate(dias_semana):
                cols_header[i].markdown(f"**{d}**")
                
            dias_marcados = []
            for semana in cal_matriz:
                cols = st.columns(7)
                for i, dia in enumerate(semana):
                    if dia == 0:
                        cols[i].write("") 
                    else:
                        data_str_atual = f"{dia:02d}/{mes_num:02d}/{ano_sel}"
                        
                        ja_marcado = False
                        desabilitar = False
                        label_caixa = str(dia)
                        
                        is_weekend = (i == 0 or i == 6)
                        is_feriado = data_str_atual in feriados_lote
                        is_conflito = data_str_atual in outros_lancamentos
                        
                        if is_conflito:
                            tipo_conflito = outros_lancamentos[data_str_atual]
                            if "Férias" in tipo_conflito:
                                label_caixa = f"{dia} (Férias)"
                            elif "Folga" in tipo_conflito:
                                label_caixa = f"{dia} (Folga)"
                            else:
                                label_caixa = f"{dia} ({tipo_conflito[:3]})"
                            desabilitar = True
                        elif is_feriado:
                            label_caixa = f"{dia} (Feriado)"
                            ja_marcado = data_str_atual in mapa_existentes
                            desabilitar = True
                        else:
                            ja_marcado = data_str_atual in mapa_existentes
                            if is_weekend:
                                desabilitar = True
                                
                        key_checkbox = f"chk_lote_{colab_id}_{dia}_{tipo_bd}_{mes_num}_{ano_sel}"
                        
                        if cols[i].checkbox(label_caixa, value=ja_marcado, disabled=desabilitar, key=key_checkbox):
                            if not is_conflito:
                                dias_marcados.append(data_str_atual)
                            
            st.markdown("<br>", unsafe_allow_html=True)
            btn_lancar_lote = st.form_submit_button(f"🚀 Gravar Lançamentos Selecionados", type="primary", use_container_width=True)
            
            if btn_lancar_lote:
                dias_marcados_set = set(dias_marcados)
                dias_existentes_set = set(mapa_existentes.keys())
                
                dias_inserir = dias_marcados_set - dias_existentes_set
                dias_remover = dias_existentes_set - dias_marcados_set
                
                if not dias_inserir and not dias_remover:
                    st.info("Nenhuma alteração foi feita nas datas.")
                else:
                    for d in dias_remover:
                        l_id, v_desc = mapa_existentes[d]
                        cursor.execute("DELETE FROM lancamentos WHERE id=%s", (l_id,))
                        if tipo_bd == "Folga BH":
                            cursor.execute("UPDATE colaboradores SET saldo_bh = saldo_bh + %s WHERE id=%s", (v_desc, colab_id))
                            
                    for d in dias_inserir:
                        cursor.execute("INSERT INTO lancamentos (colaborador_id, tipo, data_inicio, data_fim, numero_dias, valor_descontado) VALUES (%s, %s, %s, %s, %s, %s)",
                                       (colab_id, tipo_bd, d, d, 1, horas_abater))
                        if tipo_bd == "Folga BH":
                            cursor.execute("UPDATE colaboradores SET saldo_bh = saldo_bh - %s WHERE id=%s", (horas_abater, colab_id))
                            
                    st.success(f"Escala atualizada! {len(dias_inserir)} adicionados | {len(dias_remover)} removidos.")
                    st.rerun()
                    
    # ----------------------------------------------------
    # HISTÓRICO DO COLABORADOR
    # ----------------------------------------------------
    st.markdown("---")
    st.subheader("Histórico de Lançamentos")
    cursor.execute("SELECT id, tipo, data_inicio, data_fim FROM lancamentos WHERE colaborador_id=%s ORDER BY id DESC", (colab_id,))
    historico = cursor.fetchall()
    
    for h_id, t, di, df in historico:
        c1, c2 = st.columns([4, 1])
        c1.write(f"**{t}**: {di} até {df}")
        if c2.button("🗑️ Excluir", key=f"del_{h_id}"):
            cursor.execute("SELECT numero_dias, tipo, valor_descontado FROM lancamentos WHERE id=%s", (h_id,))
            res = cursor.fetchone()
            if res:
                qd, tp, vdesc = res
                if tp == "Férias (Oficial)": cursor.execute("UPDATE colaboradores SET dias_pendentes = dias_pendentes - %s WHERE id=%s", (qd, colab_id))
                elif tp == "Férias (Efetivas)": cursor.execute("UPDATE colaboradores SET dias_pendentes = dias_pendentes + %s WHERE id=%s", (qd, colab_id))
                elif tp == "Folga BH": cursor.execute("UPDATE colaboradores SET saldo_bh = saldo_bh + %s WHERE id=%s", (float(vdesc)*qd, colab_id))
                cursor.execute("DELETE FROM lancamentos WHERE id=%s", (h_id,))
                st.rerun()

# ==========================================
# MÓDULO 4: FERIADOS
# ==========================================
elif menu == "🌴 Gerir Feriados":
    st.header("Gestão de Feriados")
    
    st.subheader("🌐 Importar Feriados Nacionais (BrasilAPI)")
    st.markdown("Busca automaticamente os feriados nacionais usando a BrasilAPI e adiciona ao calendário.")
    
    with st.form("form_api"):
        col_ano, col_btn = st.columns([1, 3])
        ano_api = col_ano.number_input("Ano", value=datetime.now().year, step=1)
        
        col_btn.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
        
        submit_api = col_btn.form_submit_button("📥 Buscar da BrasilAPI", use_container_width=True)
        
        if submit_api:
            try:
                response = requests.get(f"https://brasilapi.com.br/api/feriados/v1/{ano_api}")
                if response.status_code == 200:
                    feriados_api = response.json()
                    inseridos = 0
                    for f in feriados_api:
                        d_obj = datetime.strptime(f['date'], "%Y-%m-%d")
                        d_str = d_obj.strftime("%d/%m/%Y")
                        desc = f['name']
                        
                        cursor.execute("SELECT id FROM feriados WHERE data_feriado=%s", (d_str,))
                        if not cursor.fetchone():
                            cursor.execute("INSERT INTO feriados (data_feriado, descricao) VALUES (%s, %s)", (d_str, desc))
                            inseridos += 1
                            
                    st.success(f"Sucesso! {inseridos} novos feriados importados para {ano_api}.")
                else:
                    st.error("A API não encontrou feriados para este ano ou está indisponível.")
            except Exception as e:
                st.error(f"Erro ao conectar com a BrasilAPI: {e}")
                
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        with st.form("form_feriado"):
            st.subheader("Cadastrar Feriado Manual")
            data_f = st.date_input("Data do Feriado", format="DD/MM/YYYY")
            desc_f = st.text_input("Descrição (Ex: Aniversário da Cidade)")
            if st.form_submit_button("Salvar Feriado Manual", use_container_width=True):
                try:
                    cursor.execute("INSERT INTO feriados (data_feriado, descricao) VALUES (%s, %s)", (data_f.strftime("%d/%m/%Y"), desc_f))
                    st.success("Feriado salvo!")
                    st.rerun()
                except:
                    st.error("Data já cadastrada.")
    
    with col2:
        st.subheader("Feriados Cadastrados")
        cursor.execute("SELECT id, data_feriado, descricao FROM feriados")
        feriados_banco = cursor.fetchall()
        
        feriados_banco.sort(key=lambda x: datetime.strptime(x[1], "%d/%m/%Y"), reverse=True)
        
        for fid, dt, desc in feriados_banco:
            cc1, cc2 = st.columns([4, 1])
            cc1.write(f"★ **{dt}** - {desc}")
            if cc2.button("🗑️", key=f"fdel_{fid}"):
                cursor.execute("DELETE FROM feriados WHERE id=%s", (fid,))
                st.rerun()
